#!/usr/bin/env python3
"""Mapa de armado en protoboard (hoja de cálculo, exacto por construcción).
Fuente de verdad: docs/hardware/DiseñoProtoboard.ods (geometría: ESP32 cols 22-41,
pines 25-39, USB-C col 21 F-G, libres 1-21 y 42-64) + firmware Config.h (pines).
Cada celda = un hueco real. Construido y validado zona por zona con el usuario.

Buckets (ver hoja 'Cableado y leyenda'):
 (a) EXACTO  = derivado de la fuente bloqueada + Config.h
 (b) UBICACIÓN elegida (confirmar): ULN 51-59, nodos LED, columnas ánodo/cátodo
 (c) NO VERIFICABLE sin el módulo físico: pinout DFPlayer, valor Rs
"""
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

NCOLS = 64
def fill(h): return PatternFill("solid", fgColor=h)
C_FREE=fill("FFF7EC"); C_GNDR=fill("DDDDDD"); C_3V3R=fill("FCE4C6"); C_5VR=fill("F8C9CE")
C_BODY=fill("C9CEDA"); C_VIN=fill("D7263D"); C_3V3=fill("E8821E"); C_GND=fill("333333")
C_ADC=fill("1F6FB2"); C_LED=fill("2E8B57"); C_UART=fill("8B5A2B"); C_NU=fill("E6E6E6"); C_5V=fill("D7263D")
C_USB=fill("555555"); C_RESF=fill("E7D7A6"); C_NODOT=fill("DDEBF7")
C_ULN=fill("8E7CB0"); C_DF=fill("C9B79C"); C_LEDT=fill("EDEAF6")
WHITE=Font(color="FFFFFF",bold=True,size=8); GREY=Font(color="888888",size=8)
DARK=Font(color="333333",bold=True,size=9); RESFONT=Font(color="5A4D28",bold=True,size=8)
ULNF=Font(color="FFFFFF",bold=True,size=7)
thin=Side(style="thin",color="C8C8C8"); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center"); ROT=Alignment(textRotation=90,horizontal="center",vertical="center")
BTN={1:"E63946",2:"FFD60A",3:"2A9D8F",4:"0077B6",5:"F77F00",6:"7B2CBF"}
def btnfont(i): return Font(color=("333333" if i==2 else "FFFFFF"),bold=True,size=8)

rows=["TITULO","NUMS_T","R-","R+","A","B","C","D","E","CANAL","F","G","H","I","J","R+b","R-b","NUMS_B"]
ridx={n:i+1 for i,n in enumerate(rows)}
wb=openpyxl.Workbook(); ws=wb.active; ws.title="Mapa de armado"

def cell(bc,rn): return ws.cell(row=ridx[rn],column=bc+1)
def put(bc,rn,val="",f=None,fo=None,al=CEN):
    c=cell(bc,rn); c.value=val
    if f:c.fill=f
    if fo:c.font=fo
    c.alignment=al; c.border=BORD
def lbl(c1,c2,rn,text,fillc,fo):   # etiqueta de bloque: combina celdas (evita truncado)
    ws.merge_cells(start_row=ridx[rn],start_column=c1+1,end_row=ridx[rn],end_column=c2+1)
    cc=ws.cell(ridx[rn],c1+1); cc.value=text; cc.fill=fillc; cc.font=fo; cc.alignment=CEN; cc.border=BORD

ws.cell(1,1,"TAPETE — MAPA DE ARMADO EN PROTOBOARD (830 pts) · cada celda = un hueco real · ESP32: cuerpo cols 22–41, pines 25–39 · USB-C col 21 (F–G) · ⚠ ver hoja «Cableado y leyenda»")
ws.cell(1,1).font=Font(bold=True,size=11)
for bc in range(1,NCOLS+1):
    for rn in("NUMS_T","NUMS_B"):
        c=cell(bc,rn); c.value=bc; c.font=GREY; c.alignment=CEN
for rn,txt in {"R-":"−GND","R+":"+3V3","A":"A","B":"B","C":"C","D":"D","E":"E","F":"F","G":"G","H":"H","I":"I","J":"J","R+b":"+5V","R-b":"−GND"}.items():
    c=ws.cell(ridx[rn],1,txt); c.font=DARK; c.alignment=CEN
for bc in range(1,NCOLS+1):
    for rn in("A","B","C","D","E","F","G","H","I","J"): put(bc,rn,"",C_FREE)
    put(bc,"R-","",C_GNDR); put(bc,"R+","",C_3V3R); put(bc,"R+b","",C_5VR); put(bc,"R-b","",C_GNDR)
    put(bc,"CANAL","",fill("EFE9DA"))

# ====== ZONA 1 (BLOQUEADA): ESP32 22-41, pines 25-39, USB-C 21 ======
for bc in range(22,42):
    for rn in("B","C","D","E","F","G","H","I","J"): put(bc,rn,"",C_BODY)
lbl(28,34,"D","ESP32",C_BODY,Font(bold=True,size=12)); lbl(26,37,"E","(cubre cols 22–41, filas B–J)",C_BODY,Font(size=8,italic=True))
top={25:("VIN·5V",C_VIN,WHITE),26:("GND",C_GND,WHITE),27:("D13",C_NU,GREY),28:("D12",C_NU,GREY),29:("D14",C_NU,GREY),30:("D27",C_NU,GREY),31:("D26",C_NU,GREY),32:("D25",C_NU,GREY),33:("FSR6·D33",C_ADC,WHITE),34:("FSR5·D32",C_ADC,WHITE),35:("FSR4·D35",C_ADC,WHITE),36:("FSR3·D34",C_ADC,WHITE),37:("FSR2·VN",C_ADC,WHITE),38:("FSR1·VP",C_ADC,WHITE),39:("EN",C_NU,GREY)}
for bc,(l,f,fo) in top.items(): put(bc,"A",l,f,fo,ROT)
bot={25:("3v3",C_3V3,WHITE),26:("GND",C_GND,WHITE),27:("D15",C_NU,GREY),28:("D2",C_NU,GREY),29:("LED1·D4",C_LED,WHITE),30:("DF·RX2",C_UART,WHITE),31:("DF·TX2",C_UART,WHITE),32:("LED2·D5",C_LED,WHITE),33:("LED3·D18",C_LED,WHITE),34:("LED4·D19",C_LED,WHITE),35:("LED5·D21",C_LED,WHITE),36:("RX0",C_NU,GREY),37:("TX0",C_NU,GREY),38:("D22",C_NU,GREY),39:("LED6·D23",C_LED,WHITE)}
for bc,(l,f,fo) in bot.items(): put(bc,"J",l,f,fo,ROT)
put(21,"F","USB",C_USB,WHITE); put(21,"G","-C",C_USB,WHITE)

# ====== ZONA 2 (APROBADA): 6 divisores FSR, cols 1-21 sup (3V3) ======
FSR=[(1,4,38),(2,7,37),(3,10,36),(4,13,35),(5,16,34),(6,19,33)]
fsr_jumps=[]
for i,c,adc in FSR:
    bf=fill(BTN[i])
    for rn in("D","E"): put(c,rn,"",C_NODOT)
    put(c,"A",f"B{i}",bf,btnfont(i)); put(c,"B","10k",C_RESF,RESFONT); put(c,"R-","10k",C_RESF,RESFONT)
    put(c,"C",f"B{i}▼",C_NODOT,Font(color="1F6FB2",bold=True,size=8)); put(c,"R+",f"B{i}▲",bf,btnfont(i))
    fsr_jumps.append((i,c,adc))

# ====== ZONA 3: ULN2803A (DIP-18 a caballo E/F, cols 51-59) ======
# pin1=IN1 en col51; INk=(E,50+k), OUTk=(F,50+k); GND(9)=(E,59); COM(10)=(F,59)
ULN_C0=51
for k in range(1,7):
    put(ULN_C0-1+k,"E",f"IN{k}",C_ULN,ULNF); put(ULN_C0-1+k,"F",f"OUT{k}",C_ULN,ULNF)
put(57,"E","IN7",C_NU,GREY); put(58,"E","IN8",C_NU,GREY); put(57,"F","o8",C_NU,GREY); put(58,"F","o7",C_NU,GREY)
put(59,"E","GND9",C_GND,WHITE); put(59,"F","COM10",C_5V,WHITE)
lbl(51,57,"D","ULN2803A",C_ULN,Font(color="FFFFFF",bold=True,size=9))
uln_in=[(1,29),(2,32),(3,33),(4,34),(5,35),(6,39)]   # INk(E,50+k) <- GPIO LEDk(J,col)

# ====== ZONA 4: 6 grupos LED (par ánodo/cátodo juntos, cols 1-21 inf, 5V) ======
LEDP=[(1,2,3),(2,5,6),(3,8,9),(4,11,12),(5,14,15),(6,17,18)]  # (grupo, ánodo_col, cátodo_col)
led_cath=[]
for g,ac,cc in LEDP:
    bf=fill(BTN[g])
    # ánodo: 110Ω de +5V a nodo-ánodo; hilo del tapete
    put(ac,"R+b","110",C_RESF,RESFONT); put(ac,"J","110",C_RESF,RESFONT)
    put(ac,"I","",C_LEDT); put(ac,"H",f"G{g}▲",bf,btnfont(g))
    # cátodo: hilo del tapete + jumper a OUTg
    put(cc,"F",f"G{g}",bf,btnfont(g)); put(cc,"G","",C_LEDT); put(cc,"H",f"G{g}▼",bf,btnfont(g))
    led_cath.append((g,cc,ULN_C0-1+g))   # (grupo, cátodo_col, OUT_col)

# ====== ZONA 5: DFPlayer (BLOQUE; pines PENDIENTES de pinout del módulo) ======
for c in range(42,50):
    put(c,"E","",C_DF); put(c,"F","",C_DF)
lbl(42,49,"E","DFPlayer",C_DF,Font(color="FFFFFF",bold=True,size=9))
lbl(42,49,"F","pinout PENDIENTE → hoja 2",C_DF,Font(color="FFFFFF",italic=True,size=8))

# ====== Caveats de seguridad EN EL MAPA (filas bajo la grilla) ======
def note(r,txt,color="C0392B",sz=9):
    c=ws.cell(r,1,txt); c.font=Font(bold=True,size=sz,color=color)
note(20,"⚠ SEGURIDAD — LEER ANTES DE ENERGIZAR:")
note(21,"⚠ GPIO16/RX2 NO es 5V-tolerante: mide el TX del DFPlayer (~3.3V ok; si 5V → divisor) ANTES de conectarlo a J30.")
note(22,"⚠ GPIO5/LED2 es strapping pin: si el ESP32 no arranca tras cablear LED2, reasigna LED2→GPIO22 (D22, col38) en Config.h y mueve el cable.")
note(23,"⚠ Rieles: los dos −GND se PUENTEAN (masa común); los dos + (3V3 y 5V) NUNCA se unen (cortocircuito).")
note(24,"ℹ Hoja «Cableado y leyenda»: todos los jumpers, componentes, energía y PENDIENTES (DFPlayer).","1F6FB2")

# ====== estética / page setup ======
ws.column_dimensions["A"].width=6
for bc in range(1,NCOLS+1): ws.column_dimensions[get_column_letter(bc+1)].width=4.2
for rn in("A","J"): ws.row_dimensions[ridx[rn]].height=66
for rn in("B","C","D","E","F","G","H","I"): ws.row_dimensions[ridx[rn]].height=20
ws.row_dimensions[1].height=22
ws.freeze_panes="B5"
ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1
ws.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
ws.print_area=f"A1:{get_column_letter(NCOLS+1)}24"
ws.page_margins.left=ws.page_margins.right=0.2; ws.page_margins.top=ws.page_margins.bottom=0.3

# ================= HOJA 2: cableado + leyenda + buckets =================
ws2=wb.create_sheet("Cableado y leyenda")
def r2(*v): ws2.append(list(v))
def bold(addr,sz=11): ws2[addr].font=Font(bold=True,size=sz)
r2("MAPA DE ARMADO — CABLEADO, LEYENDA Y PENDIENTES"); r2()
r2("LEYENDA DE COLORES (función)")
for t in ["rojo=VIN/5V","naranja=3V3","negro=GND","azul=ADC/FSR","verde=LED (control)","marrón=DFPlayer/UART","beige=resistencia","morado=ULN2803A"]: r2("",t)
r2()
r2("SÍMBOLOS")
r2("","Bi","nodo del FSR i (columna A–E entera = 1 punto); de aquí sale el jumper al ADC")
r2("","Bi▲ / Bi▼","hilos del tapete del FSR i: ▲ alto→+3V3, ▼ bajo→nodo")
r2("","Gi▲ / Gi▼","hilos del tapete del grupo LED i: ▲ ánodo→nodo-ánodo(110Ω), ▼ cátodo→nodo-cátodo")
r2("","10k / 110","resistencias (10k del FSR a −GND; 110Ω del grupo LED desde +5V)")
r2("","INk/OUTk","entradas/salidas del ULN2803A (canal k)")
r2()
r2("⚠ SEGURIDAD — LEER ANTES DE ENERGIZAR")
r2("","1","GPIO16/RX2 NO es 5V-tolerante: mide el TX del DFPlayer (~3.3V ok; si 5V → divisor) ANTES de conectarlo a J30.")
r2("","2","GPIO5/LED2 es strapping: si el ESP32 no arranca tras LED2, reasigna LED2→GPIO22 (D22,col38) en Config.h y mueve el cable.")
r2("","3","Rieles: los dos −GND se PUENTEAN (masa común); los dos + (3V3 y 5V) NUNCA (cortocircuito).")
r2()
r2("ENERGÍA — rieles (P1–P5)")
r2("#","Cable","Desde","Hasta","Qué hace")
for p in [("P1","3V3","ESP32 3v3 (J25)","riel superior +3V3","bus 3V3 (sensores) — el jumper cruza hacia arriba"),
          ("P2","5V","ESP32 VIN (A25)","riel inferior +5V","bus 5V (LEDs/DFPlayer) — el jumper cruza hacia abajo"),
          ("P3","GND","ESP32 GND (A26)","riel superior −GND","masa superior"),
          ("P4","GND","ESP32 GND (J26)","riel inferior −GND","masa inferior"),
          ("P5","GND","riel superior −GND","riel inferior −GND","PUENTE de masa común (1 jumper en un extremo)")]:
    r2(*p)
r2()
r2("FSR — jumpers internos (6): nodo → pin ADC del ESP32")
r2("#","Cable","Desde","Hasta","Conecta")
for i,c,adc in fsr_jumps: r2(f"F{i}",f"jumper FSR{i}",f"A{c}",f"A{adc}",f"nodo B{i} → ADC FSR{i}")
r2()
r2("LEDs — jumpers internos (12)")
r2("#","Cable","Desde","Hasta","Conecta")
for k,gp in uln_in: r2(f"Li{k}",f"IN ULN{k}",f"E{50+k}",f"J{gp}",f"IN{k} del ULN ← GPIO LED{k}")
for g,cc,oc in led_cath: r2(f"Lc{g}",f"cátodo G{g}",f"F{cc}",f"F{oc}",f"nodo-cátodo grupo {g} → OUT{g} del ULN")
r2()
r2("COMPONENTES")
r2("Comp.","Cantidad","Dónde","Nota")
r2("10kΩ","6","fila B↔−GND de cada FSR (cols 4,7,10,13,16,19)","salta el riel +3V3")
r2("110Ω","6","+5V↔fila J de cada ánodo LED (cols 2,5,8,11,14,17)","1 por grupo (decidido: 1×110Ω/grupo)")
r2("ULN2803A","1","DIP-18 a caballo E/F, cols 51–59","IN1–6=E51–56, OUT1–6=F51–56, GND9=E59→−GND, COM10=F59→+5V")
r2("DFPlayer","1","bloque cols 42–49 (E/F)","PINES PENDIENTES — ver abajo")
r2()
r2("DFPlayer — 4 conexiones (UBICAR los huecos según el pinout impreso de TU módulo)")
r2("Pin módulo","Va a","Nota")
r2("VCC","riel inferior +5V","")
r2("GND","riel inferior −GND","")
r2("RX","← Rs ← ESP32 TX2 (J31)","Rs en serie: confirmar valor (docs dicen 110Ω vs 1kΩ) — usar el del fabricante")
r2("TX","→ ESP32 RX2 (J30)","⚠ medir nivel del TX (ver seguridad #1) antes de conectar")
r2()
r2("BUCKETS DE CONFIANZA (para tu revisión)")
r2("(a) EXACTO","derivado de la fuente bloqueada + Config.h: pines del ESP32, FSR→ADC, LED→GPIO, rieles, USB-C")
r2("(b) UBICACIÓN elegida — CONFIRMAR","ULN en 51–59; columnas de nodos LED (ánodo/cátodo) y FSR; ruteo de jumpers")
r2("(c) NO VERIFICABLE sin tu módulo","pinout exacto del DFPlayer y valor de Rs (RX) — confirmar con el módulo + multímetro")
r2()
r2("CHECKLIST FÍSICO (≈10 min, próxima sesión, con el módulo y multímetro)")
for t in ["[ ] Identificar el pinout impreso del DFPlayer y fijar las celdas de VCC/GND/RX/TX en el mapa.",
          "[ ] Medir el nivel del pin TX del DFPlayer (≤3.3V directo; si 5V → divisor a GPIO16).",
          "[ ] Confirmar el valor de Rs en la línea TX2→RX del DFPlayer (110Ω vs 1kΩ).",
          "[ ] Verificar rieles: continuidad extremo a extremo, particiones, +/− no unidos (ver seguridad #3).",
          "[ ] Confirmar la colocación del ULN (51–59) y de los nodos LED en la placa real."]:
    r2("",t)
for col,w in zip("ABCDE",[14,16,20,22,60]): ws2.column_dimensions[col].width=w
ws2["A1"].font=Font(bold=True,size=13)
ws2.page_setup.orientation="landscape"; ws2.page_setup.fitToWidth=1; ws2.page_setup.fitToHeight=0
ws2.sheet_properties.pageSetUpPr=PageSetupProperties(fitToPage=True)
ws2.page_margins.left=ws2.page_margins.right=0.3

import os
out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"mapa_armado_protoboard.xlsx")
wb.save(out); print("OK ->",out)
